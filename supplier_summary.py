import openpyxl
from openpyxl.styles import Alignment
from dotenv import load_dotenv
from openai import OpenAI
import os
import re
from typing import List


def generate_supplier_summary_excel(items: List[dict], output_file: str):
    load_dotenv()
    client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suppliers"

    headers = ["S.NO", "SUPPLIER NAME", "PRICE ESTIMATE", "PACKAGING / MOQ", "WEBSITE LINK"]
    current_row = 2

    for idx, entry in enumerate(items, 1):
        item = entry["item_name"]

        prompt = f"""
You are a senior procurement officer at a construction company in Saudi Arabia.

Your job is to research and acquire high-quality materials from **real suppliers located in Saudi Arabia only**. You are preparing a procurement summary report for internal use.

### MATERIAL:
- Item: {item} Saudi Arabia

### TASK:
1. For this specific material, determine the most relevant supplier types:
   - If it is a **raw or industrial product** (e.g., steel, rebar, chemicals), focus on **local manufacturers, factories, official stockists, and authorized dealers**.
   - Avoid generic trading companies unless no direct manufacturer, dealer, or stockist is available.
2. Search ONLY for **industrial supplier websites, manufacturer product pages, official stockists, authorized local dealers** based in or shipping to Saudi Arabia and **perform an in-depth search, scour multiple pages**.
3. Return **exactly 10 Saudi-based suppliers** who offer this item or a very close variant.
4. Additionally, include **1 supplier from China** that ships internationally to Saudi Arabia.
5. For each supplier, collect:
   - Supplier name
   - Price estimate (or “Quote Required”)
   - Packaging / MOQ
   - Website link — must be a **markdown clickable link ONLY** like [Website](https://example.com)

⚠️ INSTRUCTIONS:
- Return **exactly 11 rows** per item (10 Saudi + 1 China).
- Use your judgment to select the best supplier category for the given material.
- Website link must be the supplier’s own website or direct product page (not directory or aggregator listings, unless absolutely no other link exists).
- No Amazon.
- Prefer **specialized suppliers**, **not generic shopping sites**.

### FORMAT:
Return as a markdown table with:

| # | Supplier | Price Estimate | Packaging / MOQ | Website Link |
"""

        response = client.chat.completions.create(
            model="gpt-4o-search-preview",
            messages=[{"role": "user", "content": prompt}]
        )
        markdown = response.choices[0].message.content.strip()

        print(markdown)

        # Section heading
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=6)
        ws.cell(row=current_row, column=2).value = f"MATERIAL NAME {idx}: {item}"
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center")
        current_row += 1

        # Headers
        for col, header in enumerate(headers, start=2):
            ws.cell(row=current_row, column=col).value = header
        current_row += 1

        # Parse table rows
        rows = [r.strip("|").split("|") for r in markdown.split("\n") if "|" in r][1:]

        for i, row in enumerate(rows):
            if len(row) < len(headers) - 1:
                print(f"⚠️ Skipping row {i+1}: incomplete -> {row}")
                continue

            ws.cell(row=current_row, column=2).value = i + 1  # S.NO

            for j, val in enumerate(row):
                cell_val = val.strip()
                col = 2 + j
                if col - 2 >= len(headers):
                    continue

                cell = ws.cell(row=current_row, column=col)

                if headers[col - 2] == "WEBSITE LINK":
                    match = re.search(r"\[([^\]]+)\]\((https?://[^\s)]+)\)", cell_val)
                    if match:
                        display_text = match.group(1).strip()
                        url = match.group(2).strip()
                        cell.value = display_text
                        cell.hyperlink = url
                        cell.style = "Hyperlink"
                    else:
                        match = re.search(r"(https?://[^\\s)]+)", cell_val)
                        if match:
                            url = match.group(1).strip()
                            cell.value = url
                            cell.hyperlink = url
                            cell.style = "Hyperlink"
                        else:
                            cell.value = cell_val
                else:
                    cell.value = cell_val

            current_row += 1

        current_row += 2

    wb.save(output_file)
    print(f"✅ Excel saved as: {output_file}")
