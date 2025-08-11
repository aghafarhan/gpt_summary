# summarize_doc.py
"""
Module for extracting text from PDF, DOCX, and TXT files,
summarizing it using GPT, and saving the summary to Excel.
"""
import os
import re
import pdfplumber
import pytesseract
from PIL import Image
from langdetect import detect          # currently unused → keep if using Arabic text
from docx import Document

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


from dotenv import load_dotenv
load_dotenv()
from openai import OpenAI


VERBOSE_MODE = __name__ == "__main__"  # True only if run locall

def log(*args, **kwargs):
    if VERBOSE_MODE:
        print(*args, **kwargs)


# ────────────────────────────────────────────────────────────────────────────────
# 1.  Initialise the OpenAI client
# ────────────────────────────────────────────────────────────────────────────────

client = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY")            
)
MODEL = "gpt-4.1"                                   # change to "4o" for faster usage, but may hallucinate


# ───────────────────────────────────────────────────────────────────
# 2.  Helpers to extract raw-text *and* run the description normaliser
# ───────────────────────────────────────────────────────────────────

CODE_RE   = re.compile(r"^[A-Z0-9\-\/]+$")   # first token looks like a code
BRAND_RE  = re.compile(r"^[A-Z]{3,}$")       # second token looks like BRAND


ADJECTIVES = {"SHUTTERING", "MARINE", "ORD", "INDONESIAN", "SMARTPLEX"}

# Generic numeric pair   4*8  4-8  4×8  4 X 8  →  4X8
DIM_RE = re.compile(
    r"\b(\d+)\s*[×xX*–—-]\s*(\d+)\b",         # accepts hyphen & long dash
    flags=re.UNICODE,
)


# Millimetre pair        150MM-12MM / 150MM*12 / 150 MM – 12 mm → 150MMX12MM
MM_DIM_RE = re.compile(
    r"\b(\d+)MM\s*[×xX*–—-]\s*(\d+)(?:MM)?\b",   # 2nd “MM” optional
    flags=re.UNICODE | re.IGNORECASE,
)

MM_RE    = re.compile(r"\b(\d+)\s*MM\b", re.I)                # 18 mm → 18MM


LONG_DIM_RE = re.compile(
    r"\b\d+(?:\s*[×xX*–—-]\s*\d+){2,}\b",          # ≥2 repetitions
    flags=re.UNICODE,
)

def _fold_long_dims(m: re.Match) -> str:
    """'18 * 122 * 244' → '18X122X244'."""
    nums = re.findall(r"\d+", m.group(0))
    return "X".join(nums)

def normalise_description(line: str) -> str:
    """(unchanged – keeps human wording)"""
    tokens = line.strip().split()
    if tokens and CODE_RE.match(tokens[0]):
        tokens = tokens[1:]
        if tokens and BRAND_RE.match(tokens[0]):
            tokens = tokens[1:]
    desc = " ".join(tokens)
    return re.sub(r"\s{2,}", " ", desc).strip().upper()


def canonical_key(desc: str) -> str:
    desc = desc.upper()
    desc = re.sub(r"[,/]+", " ", desc)

    # 1️⃣ 18 mm  → 18MM
    desc = MM_RE.sub(lambda m: f"{m.group(1)}MM", desc)

    # 2️⃣-a generic 4×8 pair
    desc = DIM_RE.sub(lambda m: f"{m.group(1)}X{m.group(2)}", desc)

    # 2️⃣-b long sequences 18*122*244
    desc = LONG_DIM_RE.sub(_fold_long_dims, desc)

    # 2️⃣-c 150MM-12MM pair
    desc = MM_DIM_RE.sub(lambda m: f"{m.group(1)}MMX{m.group(2)}MM", desc)

    # 3️⃣ drop filler adjectives / brands
    tokens = [t for t in desc.split()
              if t not in ADJECTIVES and not BRAND_RE.match(t)]
    return " ".join(tokens).strip()



def extract_text_from_pdf(path: str) -> str:
    """
    • Try to read the PDF’s text layer with pdfplumber.
    • If that returns nothing on every page, fall back to OCR.
    • Every line (from either route) is piped through `normalise_description`
      so the downstream prompt sees tidy material descriptions.
    """
    import pdfplumber, pytesseract
    from PIL import Image

    out = []

    with pdfplumber.open(path) as pdf:
        had_text = False
        for page in pdf.pages:
            raw = page.extract_text() or ""
            if raw.strip():
                had_text = True
            for ln in raw.splitlines():
                pretty = normalise_description(ln)
                if not pretty:
                    continue
                key = canonical_key(pretty)
                out.append(f"{key} | {pretty}")

    return "\n".join(out)


# ────────────────────────────────────────────────────────────────────────────────
# 3.  Prompt & call GPT
# ────────────────────────────────────────────────────────────────────────────────

def summarize_text_with_gpt(combined_text: str) -> str:
    """
    Feed the concatenated raw text from all quotation files to GPT and
    receive two markdown tables:
       1) Material & Supplier Comparison  (Unit-price only, no totals)
       2) Payment Terms & Quotation Validity
    """
    prompt = f"""
You are an intelligent business assistant specialised in supplier quotations.

------------------------------------------------------------------------------
### 1  📦 Material & Supplier Comparison (Unit Price only)
Canonical Key (CK) ✨  
• Before tabulating, treat the **left-hand token** (everything before the first
  “|”) as the CK.  It is already normalised in Python so, for example,  
  `150MM-12MM`, `150 MM – 12 MM`, `150MM*12MM` ⇒ **CK `150MMX12MM`**  
  and `4 * 8`, `4×8`, `4-8` ⇒ **CK `4X8`** and `244 * 122 * 18` ⇒ **CK `244X122X18`**.  
• Every distinct CK is **one row**.  Do **not** merge rows with different CKs.

Table rules  
• Columns **exactly**:  
  `SN | Altered Material Name | Material description | <Supplier-1> Unit Price | <Supplier-2> Unit Price | … | Source file(s)`  
• If a supplier did **not** quote that item → Unit Price `N/A`.  
• No totals, no VAT, no grand totals, no commentary.  
• Output valid markdown.

### 2  💳 Payment Terms & Quotation Validity
Columns: `Supplier Name | Payment Terms | Quotation Validity`  
Copy wording exactly; leave blank if absent.


--- BEGIN DOCUMENTS ---
{combined_text}
--- END DOCUMENTS ---
"""
    response = client.chat.completions.create(
        model=MODEL,
        messages=[
            {
                "role": "system",
                "content": 
                (
                "You are a document-summariser. Your job is to summarise supplier "
                "quotations into markdown tables. For every material in a row, list the unit "
                "price/Rate quoted by each supplier across the documents. Also make sure that all the materials is listed even if they are not comparable. No extra prose—"
                "output only the required tables."
                )
            },
            {"role": "user", "content": prompt},
        ],
    )
    return response.choices[0].message.content


# ────────────────────────────────────────────────────────────────────────────────
# 4.  Markdown  →  DataFrame (robust version, keeps empty cells)
# ────────────────────────────────────────────────────────────────────────────────

def markdown_table_to_df(md_block: str) -> pd.DataFrame:
    """
    Convert a pipe-delimited markdown table to DataFrame **without**
    discarding empty cells.  Leading / trailing pipes are trimmed once so
    column counts remain stable.
    """
    lines = [ln.rstrip() for ln in md_block.splitlines() if "|" in ln]

    if len(lines) < 2:
        raise ValueError("No markdown table found.")

    def split_row(row: str) -> list[str]:
        row = row.strip()
        if row.startswith("|"):
            row = row[1:]
        if row.endswith("|"):
            row = row[:-1]
        return [c.strip() for c in row.split("|")]

    header = split_row(lines[0])
    # Skip the separator row (---|---) which is always the 2nd line
    body   = [split_row(r) for r in lines[2:] if r.strip()]

    # Pad rows that are shorter than header with empty strings
    fixed_body = [r + [""] * (len(header) - len(r)) if len(r) < len(header) else r
                  for r in body]

    return pd.DataFrame(fixed_body, columns=header)


# ────────────────────────────────────────────────────────────────────────────────
# 5.  Save tables to Excel (uses the robust parser & never closes empty wb)
# ────────────────────────────────────────────────────────────────────────────────

def save_summary_to_excel(summary_text: str,
                          output_path: str = "combined_summary.xlsx") -> None:

    blocks = [b for b in summary_text.split("\n\n") if "|" in b and "-" in b]
    if not blocks:
        log("⚠️  GPT returned no markdown tables.")
        return

    start_row = 0
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for i, block in enumerate(blocks):
            try:
                df = markdown_table_to_df(block)
            except Exception as e:
                log(f"⚠️  Skipping one table – parse error: {e}")
                continue

            df.to_excel(writer, sheet_name="Summary",
                        index=False, startrow=start_row)

            # Grand-total logic only for the first (materials) table
            if i == 0:
                total_row = [""] * len(df.columns)
                total_row[0] = "Grand Total"
                for idx, col in enumerate(df.columns):
                    if "Total" in col:
                        nums = pd.to_numeric(df[col], errors="coerce")
                        sm   = nums.dropna().sum()
                        total_row[idx] = round(sm, 2) if sm else ""
                pd.DataFrame([total_row], columns=df.columns).to_excel(
                    writer, sheet_name="Summary",
                    index=False, startrow=start_row + len(df) + 1
                )
                start_row += len(df) + 5
            else:
                start_row += len(df) + 4

    # Re-open for colouring – guarantee at least one sheet exists
    wb = load_workbook(output_path)
    ws = wb["Summary"]

    green  = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFF9C4")

    # low-price highlight for first table only
    end_row = next((r for r in range(2, ws.max_row + 1)
                    if all((ws.cell(r, c).value in (None, "") for c in range(1, ws.max_column + 1)))), ws.max_row)

    for r in ws.iter_rows(min_row=2, max_row=end_row - 1):
        numeric = [(c, float(c.value))
                   for c in r[2:]
                   if isinstance(c.value, (int, float, str))
                   and str(c.value).replace('.', '', 1).isdigit()]
        if numeric:
            low = min(v for _, v in numeric)
            for cell, val in numeric:
                if val == low:
                    cell.fill = green

    # yellow grand-total row (last row)
    for cell in ws[ws.max_row]:
        cell.fill = yellow

    wb.save(output_path)
    log(f"✅ Excel summary saved → {output_path}")


# ───────────────────────────────────────────────────────────────────────────────────────────
# 6.  Run_local_test: read all files in folder, send to GPT, save Excel (NOT USED IN BACKEND)
# ───────────────────────────────────────────────────────────────────────────────────────────

def run_local_test():
    folder = input("📁 Folder with quotation files: ").strip()
    if not os.path.isdir(folder):
        log("❌ Folder not found.")
        return

    EXT = {".pdf", ".docx", ".txt"}
    files = [f for f in os.listdir(folder) if os.path.splitext(f)[1].lower() in EXT]
    if not files:
        log("⚠️ No supported files in folder.")
        return

    log(f"✅ Found {len(files)} file(s):", *files, sep="\n  • ")

    combined = ""
    for f in files:
        path = os.path.join(folder, f)
        log(f"   ─ reading {f}")
        try:
            ext = os.path.splitext(f)[1].lower()
            if ext == ".pdf":
                txt = extract_text_from_pdf(path) 
                
            combined += f"\n\n--- FILE: {f} ---\n\n{txt}"
        except Exception as exc:
            log(f"     ⚠️ {f}: {exc}")

    if not combined.strip():
        log("❌ No text extracted.")
        return

    log("🧠 Sending to GPT …")
    md_summary = summarize_text_with_gpt(combined)
    log("✅ GPT returned markdown tables\n")

    log(md_summary)               # show in console
    save_summary_to_excel(md_summary)


if __name__ == "__main__":

    run_local_test()


